import tkinter as tk
from tkinter import ttk
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Alignment
from PIL import Image, ImageTk
from openpyxl import load_workbook
import tkinter as tk
from tkinter import messagebox
from tkinter import filedialog, messagebox
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import os
from reportlab.pdfgen import canvas
from datetime import datetime
import datetime
import modulefinder
from datetime import datetime
import requests
import io
from io import BytesIO
import win32com.client as win32
import pandas as pd
import shutil
import openpyxl
import calendar
import win32api
import os
import time
from reportlab.pdfgen import canvas

janela = tk.Tk()
janela.title("Sistema de Pedidos")
janela.geometry("600x400")

# Fazer o download da imagem a partir do URL
url = "https://i.imgur.com/FQODni0.jpeg"
resposta = requests.get(url)
imagem = Image.open(BytesIO(resposta.content))

# Redimensionar a imagem
imagem_redimensionada = imagem.resize((100, 100))

# Criar a PhotoImage
icone_tk = ImageTk.PhotoImage(imagem_redimensionada)

# Definir o ícone da janela
janela.iconphoto(True, icone_tk)

# Referência global para a PhotoImage
janela.icone_tk = icone_tk

data = date.today().strftime("%d/%m/%Y")
data_label = tk.Label(janela, text=data, font=("Arial", 14))
data_label.pack(pady=20)

url_imagem = "https://i.imgur.com/FQODni0.jpeg"
resposta = requests.get(url_imagem)
imagem = Image.open(BytesIO(resposta.content))
imagem = imagem.resize((100, 100))
imagem_tk = ImageTk.PhotoImage(imagem)

label_imagem = tk.Label(janela, image=imagem_tk)
label_imagem.pack()

# Criar uma planilha do Excel

# Variável para controlar o número da linha na planilha
linha_atual = 2

# Lista de produtos
produtos_só_acaí = [
    {'tipo': 'Açaí 296 ML', 'valor': 15},
    {'tipo': 'Açaí 473 ML', 'valor': 18},
    {'tipo': 'Açaí 710 ML', 'valor': 21},
]

produtos_no_pote = [
    {'tipo': 'Açaí no pote - 296 ML', 'valor': 18},
    {'tipo': 'Açaí no pote - 473 ML', 'valor': 22},
    {'tipo': 'Açaí no pote - 710 ML', 'valor': 25},

]

produtos_bebidas = [
    {'tipo': 'Água sem gás', 'valor': 3},
    {'tipo': 'Águas com gás', 'valor': 4},
    {'tipo': 'Suco', 'valor': 6},
    {'tipo': 'Refrigerante', 'valor': 4},
]

adicionais = [
    {'tipo': 'Leite em pó', 'valor': 2},
    {'tipo': 'Granola', 'valor': 2},
    {'tipo': 'Paçoca', 'valor': 2},
    {'tipo': 'Leite condensado', 'valor': 2},
    {'tipo': 'Flocos de arroz', 'valor': 2},
    {'tipo': 'Flocos de tapioca', 'valor': 2},
    {'tipo': 'Cobertura de morango', 'valor': 2},
    {'tipo': 'Cobertura de chocolate', 'valor': 2},
    {'tipo': 'Sucrilhos', 'valor': 2},
    {'tipo': 'Amendoim', 'valor': 2},
]

complementos = [
    {'tipo': 'Morango', 'valor': 3},
    {'tipo': 'Uva', 'valor': 3},
    {'tipo': 'Banana', 'valor': 3},
    {'tipo': 'Creme ninho', 'valor': 3},
    {'tipo': 'Creme nutella', 'valor': 3},
    {'tipo': 'Creme oreo', 'valor': 3},
    {'tipo': 'Creme morango', 'valor': 3},
    {'tipo': 'Confete', 'valor': 3},
    {'tipo': 'Ovomaltine', 'valor': 3},
    {'tipo': 'chocoball', 'valor': 3},
    {'tipo': 'Farinha Láctea', 'valor': 3},
    {'tipo': 'Mel', 'valor': 3},
]

adicionais2 = [
    {'tipo': 'Leite em pó', 'valor': 0},
    {'tipo': 'Granola', 'valor': 0},
    {'tipo': 'Paçoca', 'valor': 0},
    {'tipo': 'Leite condensado', 'valor': 0},
    {'tipo': 'Flocos de arroz', 'valor': 0},
    {'tipo': 'Flocos de tapioca', 'valor': 0},
    {'tipo': 'Cobertura de morango', 'valor': 0},
    {'tipo': 'Cobertura de chocolate', 'valor': 0},
    {'tipo': 'Sucrilhos', 'valor': 0},
    {'tipo': 'Amendoim', 'valor': 0},
]

complementos2 = [
    {'tipo': 'Morango', 'valor': 0},
    {'tipo': 'Uva', 'valor': 0},
    {'tipo': 'Banana', 'valor': 0},
    {'tipo': 'Creme ninho', 'valor': 0},
    {'tipo': 'Creme nutella', 'valor': 0},
    {'tipo': 'Creme oreo', 'valor': 0},
    {'tipo': 'Creme morango', 'valor': 0},
    {'tipo': 'Confete', 'valor': 0},
    {'tipo': 'Ovomaltine', 'valor': 0},
    {'tipo': 'chocoball', 'valor': 0},
    {'tipo': 'Farinha Láctea', 'valor': 0},
    {'tipo': 'Mel', 'valor': 0},
]

mesas = []  # Lista para armazenar as mesas e seus pedidos
vendas = []  # Lista para armazenar as vendas
pedido = vendas
valor_pix = []
valor_debito = []
valor_credito = []
valor_dindin = []
caixa_inicial = 0.0  # Valor do caixa inicial
# PASTA
import os

pasta_arquivos = os.path.join(os.path.expanduser("~"), "Documents", "Extremosabor")

# Verifica se a pasta já existe
if not os.path.exists(pasta_arquivos):
    # Cria a pasta
    os.makedirs(pasta_arquivos)
    print("Pasta 'ExtremoSabor' criada com sucesso!")
else:
    print("A pasta 'Extremosabor' já existe.")

# PASTA DOS RELATÓRIOS
pasta_relatorio = os.path.join(os.path.expanduser("~"), "Documents", "Extremosabor", "Relatório")

# Verifica se a pasta já existe
if not os.path.exists(pasta_relatorio):
    # Cria a pasta
    os.makedirs(pasta_relatorio)
    print("Pasta 'ExtremoSabor' criada com sucesso!")
else:
    print("A pasta 'Extremosabor' já existe.")

# PLANILHA
pasta_documentos = os.path.join(os.path.expanduser("~"), "Documents", "ExtremoSabor", "Relatório")
nome_pasta = 'Relatório-' + date.today().strftime("%Y-%h")
pasta_mes_atual = os.path.join(pasta_documentos, nome_pasta)
data_resetada = date.today().strftime("%d-%m-%Y")
nome_arquivo = f"Relatorio_{data_resetada}.xlsx"
caminho_arquivo = os.path.join(pasta_documentos, pasta_mes_atual, nome_arquivo)
if not os.path.exists(pasta_mes_atual):
    # Create the folder for the current month if it doesn't exist
    os.makedirs(pasta_mes_atual)
    print(f"Pasta '{nome_pasta}' criada com sucesso!")
else:
    print(f"A pasta '{nome_pasta}' já existe.")

if not os.path.exists(caminho_arquivo):
    # Cria um novo arquivo se não existir
    workbook = Workbook()
    workbook.save(filename=caminho_arquivo)

workbook = load_workbook(filename=caminho_arquivo)
planilha_ativa = workbook.active

planilha_ativa['A1'] = "Item"
planilha_ativa['B1'] = "Quantidade"
planilha_ativa['C1'] = "Valor unitário"
planilha_ativa['D1'] = "Valor Total"
# Somar os valores da coluna D
total_diario = sum([cell.value for cell in planilha_ativa['D'] if isinstance(cell.value, (int, float))])

# Escrever o resultado na célula E2
planilha_ativa['E2'] = total_diario

# Escrever "Total diário" na célula E1
planilha_ativa['E1'] = "Total diário"

# Salvar as alterações na planilha
workbook.save(filename=caminho_arquivo)


def adicionar_pedido_planilha(tipo_produto, valor_produto):
    global linha_atual
    pedidos_existentes = set()
    for row in planilha_ativa.iter_rows(min_row=2, max_row=linha_atual - 1, min_col=1, max_col=1):
        for cell in row:
            pedidos_existentes.add(cell.value)

    if tipo_produto in pedidos_existentes:
        for row in planilha_ativa.iter_rows(min_row=2, max_row=linha_atual - 1, min_col=1, max_col=4):
            for cell in row:
                if cell.value == tipo_produto:
                    quantidade_atual = planilha_ativa.cell(row=cell.row, column=2).value
                    valor_atual = planilha_ativa.cell(row=cell.row, column=3).value
                    if valor_atual is None:
                        valor_atual = 0
                    planilha_ativa.cell(row=cell.row, column=2).value = quantidade_atual + 1
                    planilha_ativa.cell(row=cell.row, column=4).value = (quantidade_atual + 1) * valor_atual
                    break
    else:
        nova_linha = [tipo_produto, 1, valor_produto, valor_produto]
        planilha_ativa.append(nova_linha)
        linha_atual += 1

    mesa['valor_total'] += valor_produto
    atualizar_tabela_mesas()
    workbook.save(filename=caminho_arquivo)

def planilha_diaria_atualizar_valores():
    # Carregar a planilha existente
    planilha = openpyxl.load_workbook(filename=caminho_arquivo)
    planilha_ativa = planilha.active

    # Calcular o novo valor total
    valor_total = sum([cell.value for cell in planilha_ativa['D'] if isinstance(cell.value, (int, float))])

    # Atualizar o valor total na célula correspondente
    planilha_ativa['E2'] = valor_total

    # Salvar as alterações na planilha
    planilha.save(filename=caminho_arquivo)
    planilha.close()
# Cupom
def enviar_arquivos_por_email(diretorio, destinatario):
    data_resetada = date.today().strftime("%d-%m-%Y")
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = destinatario
    mail.Subject = f'Cupons (Qe 30) do dia {data_resetada}'
    mail.HTMLBody = f'''
    <p>Prezado, Rafael</p>

    <p>Seguem os cupons do dia {data_resetada}.</p>

    <p>Att.,</p>
    <p>Extremo sabor.</p>
    '''

    for nome_arquivo in os.listdir(diretorio):
        caminho_arquivo = os.path.join(diretorio, nome_arquivo)
        if os.path.isfile(caminho_arquivo):  # Verifica se é um arquivo
            attachment = os.path.abspath(caminho_arquivo)
            mail.Attachments.Add(attachment)

    mail.Send()
    print('Email enviado com sucesso.')


contador_cliente = 1


def salvar_relatorio_pdf():
    global contador_cliente
    global diretorio2

    # Diretório base para salvar o relatório
    diretorio_base = os.path.join(os.path.expanduser("~"), "Documents", "ExtremoSabor", "Nota Fiscal")

    # Obter a data atual
    data_atual = datetime.date.today()
    nome_pasta = 'Nota Fiscal-' + data_atual.strftime("%Y-%m")

    # Diretório completo para salvar o relatório
    diretorio = os.path.join(diretorio_base, nome_pasta)
    nome_pasta2 = data_atual.strftime("%d-%m-%Y")
    diretorio2 = os.path.join(diretorio_base, nome_pasta, nome_pasta2)

    # Verificar se o diretório existe, caso contrário, criar o diretório
    if not os.path.exists(diretorio2):
        os.makedirs(diretorio2)

    # Nome do arquivo com número do cliente e data
    nome_arquivo = f"Cupom{contador_cliente}-{data_atual.strftime('%d-%m-%Y')}.pdf"
    caminho_arquivo = os.path.join(diretorio2, nome_arquivo)

    # Verificar se o arquivo para a mesa 1 já existe
    if contador_cliente == 1 and os.path.exists(caminho_arquivo):
        contador_cliente += 1
        nome_arquivo = f"Cupom{contador_cliente}-{data_atual.strftime('%d-%m-%Y')}.pdf"
        caminho_arquivo = os.path.join(diretorio2, nome_arquivo)

    # Atualizar o contador do cliente
    contador_cliente += 1

    c = canvas.Canvas(caminho_arquivo)
    c.setFont("Helvetica-Bold", 10)

    # Escrever o cabeçalho do relatório
    c.drawString(30, 780, f"Açaí extremo sabor")
    c.drawString(30, 765, "Pedidos:")

    # Escrever os dados dos pedidos na tabela de relatório
    y = 750
    tabela_mesas.delete(*tabela_mesas.get_children())
    valor_total = 0  # Inicializa o valor total como zero
    for pedido in mesa['pedidos']:
        tipo_pedido = pedido['tipo']
        valor_pedido = pedido['valor']
        pagamento = mesa['forma_pagamento']
        c.drawString(30, y, f"{tipo_pedido}")
        c.drawString(160, y, f" - R$ {valor_pedido:.2f}")
        y -= 20
        valor_total += valor_pedido  # Adiciona o valor do pedido ao valor total

    # Pular uma linha antes de escrever o valor total
    y -= 20

    # Escrever o valor total
    c.drawString(30, y, f"Valor Total: R$ {valor_total:.2f}")

    # Fechar o PDF
    c.save()

    #Impressão
    win32api.ShellExecute(0, "print", caminho_arquivo, None, ".", 0)








# PASTA DOS RELATÓRIOS MENSAIS
pasta_relatorio = os.path.join(os.path.expanduser("~"), "Documents", "Extremosabor", "Relatório mensal")

# Verifica se a pasta já existe
if not os.path.exists(pasta_relatorio):
    # Cria a pasta
    os.makedirs(pasta_relatorio)
    print("Pasta 'Relatório mensal' criada com sucesso!")
else:
    print("A pasta 'Relatório mensal' já existe.")
# Obtém o mês e o ano atual
mes_ano_atual = datetime.now().strftime("%B %Y")

# Cria o caminho da pasta dentro de "Relatório mensal"
pasta_mes_ano = os.path.join(pasta_relatorio, mes_ano_atual)

# Verifica se a pasta já existe
if not os.path.exists(pasta_mes_ano):
    # Cria a pasta
    os.makedirs(pasta_mes_ano)
    print(f"Pasta '{mes_ano_atual}' criada com sucesso dentro de 'Relatório mensal'!")
else:
    print(f"A pasta '{mes_ano_atual}' já existe dentro de 'Relatório mensal'.")


def criar_planilha_menu(itens_menu):
    # Cria um novo arquivo de planilha
    workbook = Workbook()
    planilha = workbook.active

    # Cabeçalho da planilha
    planilha['A1'] = "Item"
    planilha['B1'] = "Quantidade"
    planilha['C1'] = "Valor Unitário"
    planilha['D1'] = "Valor Total"

    # Preenche a planilha com os itens do menu
    linha_atual = 2
    for item in itens_menu:
        planilha.cell(row=linha_atual, column=1, value=item['tipo'])
        planilha.cell(row=linha_atual, column=2, value=0)
        planilha.cell(row=linha_atual, column=3, value=item['valor'])

        linha_atual += 1

    # Salva a planilha em um arquivo
    data_resetada2 = date.today().strftime("%m-%Y")
    nome_arquivo = f"Relatório Mensal{data_resetada2}.xlsx"

    caminho_arquivo = os.path.join(pasta_mes_ano, nome_arquivo)
    if not os.path.exists(caminho_arquivo):
        # Cria a pasta
        workbook.save(filename=caminho_arquivo)
        print(f"Planilha 'Relatório Mensal-{data_resetada2}' criada com sucesso!")
    else:
        print(f"A planilha 'Relatório Mensal-{data_resetada2}' já existe.")


# Lista com todos os produtos do menu
itens_menu = (
        produtos_só_acaí +
        produtos_no_pote +
        produtos_bebidas +
        adicionais +
        complementos +
        adicionais2 +
        complementos2
)

# Chama a função para criar a planilha do menu
criar_planilha_menu(itens_menu)


def adicionar_pedido3(tipo_produto, valor_produto):
    # Atualiza a quantidade do item na lista de pedidos da mesa
    mesa['pedidos'].append({'tipo': tipo_produto, 'valor': valor_produto})
    mesa['valor_total'] += valor_produto

    # Atualiza a quantidade do item no arquivo menu.xlsx
    arquivo_menu = "Relatório Mensal{}.xlsx".format(date.today().strftime("%m-%Y"))
    caminho_arquivo = os.path.join(pasta_relatorio, arquivo_menu)

    workbook = openpyxl.load_workbook(caminho_arquivo)
    planilha = workbook.active

    linha_atual = 2
    for row in planilha.iter_rows(min_row=linha_atual):
        if row[0].value == tipo_produto:
            row[1].value = row[1].value + 1
            break
        linha_atual += 1

    workbook.save(caminho_arquivo)


def pix():
    total_pix = caixa_inicial
    for mesa in vendas:
        forma_pagamento = mesa['forma_pagamento']
        if forma_pagamento == 'Pix':
            valor_venda = mesa['valor_total']
            total_pix += valor_venda / 2
    valor_pix.append(total_pix)

    print(sum(valor_pix))
    return total_pix


def credito():
    total_credito = caixa_inicial
    for mesa in vendas:
        forma_pagamento = mesa['forma_pagamento']
        if forma_pagamento == 'Crédito':
            valor_venda = mesa['valor_total']
            total_credito += valor_venda / 2

    valor_credito.append(total_credito)
    print(sum(valor_credito))
    return total_credito


def debito():
    total_debito = caixa_inicial
    for mesa in vendas:
        forma_pagamento = mesa['forma_pagamento']
        if forma_pagamento == 'Débito':
            valor_venda = mesa['valor_total']
            total_debito += valor_venda / 2

    valor_debito.append(total_debito)
    print(sum(valor_debito))
    return total_debito


# FINALIZAR MESA
def finalizar_mesa():
    global tipo_pedido
    forma_pagamento = forma_pagamento_combobox.get()
    if forma_pagamento:
        mesa['forma_pagamento'] = forma_pagamento
        mesas.append(mesa)
        vendas.append(mesa)  # Adiciona a mesa à lista de vendas
        atualizar_tabela_mesas()
        atualizar_total_caixa()  # Atualiza o total do caixa
        salvar_relatorio_pdf()
        janela_mesa.destroy()
        # Salvar a planilha do Excel
        nome_arquivo = f"Relatorio_{date.today().strftime('%Y%m%d_%H%M%S')}.xlsx"


def atualizar_tabela_mesas():
    tabela_mesas.delete(*tabela_mesas.get_children())
    for pedido in mesa['pedidos']:
        tabela_mesas.insert("", "end", values=(pedido['tipo'], f"R$ {pedido['valor']:.2f}"))


def atualizar_total_caixa():
    total_caixa = caixa_inicial + sum(mesa['valor_total'] for mesa in vendas)
    total_caixa_label.config(text=f"Total: R$ {total_caixa / 2:.2f}")


# CAIXAM

def adicionar_pedido2(tipo_produto, valor_produto):
    mesa['pedidos'].append({'tipo': tipo_produto, 'valor': valor_produto})
    mesa['valor_total'] += valor_produto
    atualizar_tabela_mesas()
    adicionar_pedido_planilha(tipo_produto, valor_produto)
    # Atualiza a quantidade do item no arquivo menu.xlsx
    arquivo_menu = "Relatório Mensal{}.xlsx".format(date.today().strftime("%m-%Y"))
    caminho_arquivo = os.path.join(pasta_mes_ano, arquivo_menu)
    if 'Açaí no pote' in tipo_produto:
        abrir_janela_adicionais_pote()
        if tipo_produto == 'Açaí no pote - 473 ML' or tipo_produto == 'Açaí no pote - 710 ML':
            abrir_janela_complementos_pote()
    workbook = openpyxl.load_workbook(caminho_arquivo)
    planilha = workbook.active

    linha_atual = 2
    for row in planilha.iter_rows(min_row=linha_atual):
        if row[0].value == tipo_produto:
            quantidade = row[1].value + 1
            row[1].value = quantidade
            valor_unitario = row[2].value
            valor_total = quantidade * valor_unitario
            row[3].value = valor_total
            break
        linha_atual += 1

    workbook.save(caminho_arquivo)

janela_complementos_pote_aberta = False
janela_complementos_pote = None

def abrir_janela_adicionais_pote():
    janela_adicionais_pote = tk.Toplevel(janela_mesa)
    janela_adicionais_pote.title("Adicionais Pote")
    janela_adicionais_pote.geometry("+{}+{}".format(janela_mesa.winfo_x() + janela_mesa.winfo_width(), janela_mesa.winfo_y()))

    for produto in adicionais2:
        tipo_produto = produto['tipo']
        valor_produto = produto['valor']

        botao_produto = tk.Button(janela_adicionais_pote, text=f"{tipo_produto}",
                                  command=lambda tipo_produto=tipo_produto, valor_produto=valor_produto:
                                  adicionar_pedido2(tipo_produto, valor_produto))
        botao_produto.pack()

    botao_fechar = tk.Button(janela_adicionais_pote, text="Fechar",font=("Arial", 14, "bold"), command=janela_adicionais_pote.destroy)
    botao_fechar.pack()

def abrir_janela_complementos_pote():
    global janela_complementos_pote_aberta, janela_complementos_pote

    if not janela_complementos_pote_aberta:
        deslocamento_vertical = janela_mesa.winfo_height()
        janela_complementos_pote = tk.Toplevel(janela_mesa)
        janela_complementos_pote.title("Complementos Pote")
        janela_complementos_pote.geometry("+{}+{}".format(janela_mesa.winfo_x() + janela_mesa.winfo_width(), janela_mesa.winfo_y() + deslocamento_vertical))

        for produto in complementos2:
            tipo_produto = produto['tipo']
            valor_produto = produto['valor']

            botao_produto = tk.Button(janela_complementos_pote, text=f"{tipo_produto}",
                                      command=lambda tipo_produto=tipo_produto, valor_produto=valor_produto:
                                      adicionar_pedido2(tipo_produto, valor_produto))
            botao_produto.pack()

        botao_fechar = tk.Button(janela_complementos_pote, text="Fechar",font=("Arial", 14, "bold"), command=fechar_janela_complementos_pote)
        botao_fechar.pack()

        janela_complementos_pote.protocol("WM_DELETE_WINDOW", fechar_janela_complementos_pote)
        janela_complementos_pote_aberta = True

def fechar_janela_complementos_pote():
    global janela_complementos_pote_aberta, janela_complementos_pote
    janela_complementos_pote_aberta = False
    janela_complementos_pote.destroy()

def abrir_janela_mesa():
    global mesa, forma_pagamento_combobox, total_caixa_label, tabela_mesas, janela_mesa
    mesa = {'pedidos': [], 'valor_total': 0, 'forma_pagamento': ''}  # Dicionário para representar a mesa

    janela_mesa = tk.Toplevel(janela)
    janela_mesa.title("Mesa")

    scrollbar = tk.Scrollbar(janela_mesa)
    scrollbar.pack(side="right", fill="y")

    # Crie um frame principal para conter todos os widgets
    principal_frame = ttk.Frame(janela_mesa)

    tabela_mesas_frame = tk.Frame(principal_frame)
    tabela_mesas_frame.pack(padx=0, pady=10)

    tabela_mesas = ttk.Treeview(tabela_mesas_frame, columns=(1, 2), show="headings", height=15)
    tabela_mesas.pack(side="left")

    tabela_mesas.heading(1, text="Pedido")
    tabela_mesas.heading(2, text="Valor")

    scrollbar.configure(command=tabela_mesas.yview)
    tabela_mesas.configure(yscrollcommand=scrollbar.set)

    atualizar_tabela_mesas()

    total_caixa_label = tk.Label(principal_frame, text="Total: R$ 0.00", font=("Arial", 12))
    total_caixa_label.pack(pady=20)

    forma_pagamento_label = tk.Label(principal_frame, text="Pagamento:", font=("Arial", 12))
    forma_pagamento_label.pack()

    forma_pagamento_combobox = ttk.Combobox(principal_frame, values=["Dinheiro", "Débito", 'Crédito', "Pix"])
    forma_pagamento_combobox.pack()

    # TESTE CUPOM

    finalizar_mesa_button = tk.Button(janela_mesa, text="Finalizar Mesa", command=finalizar_mesa)
    finalizar_mesa_button.pack(padx=1, pady=20, side='bottom')

    atualizar_total_caixa()
    # Só ACAI
    produtos_só_acaí_frame = tk.Frame(janela_mesa)
    produtos_só_acaí_frame.pack(pady=20, side="left", anchor="n")

    produtos_só_acaí_label = tk.Label(produtos_só_acaí_frame, text="Só Açaí", font=("Arial", 14))
    produtos_só_acaí_label.pack()

    for produto in produtos_só_acaí:
        tipo_produto = produto['tipo']
        valor_produto = produto['valor']

        botão_produto = tk.Button(produtos_só_acaí_frame, text=f"{tipo_produto} - R$ {valor_produto:.2f}",
                                  command=lambda tipo_produto=tipo_produto, valor_produto=valor_produto:
                                  adicionar_pedido2(tipo_produto, valor_produto))
        botão_produto.pack()

    # ACAI NO POTE
    produtos_no_pote_frame = tk.Frame(janela_mesa)
    produtos_no_pote_frame.pack(pady=20, side="left", anchor="n")

    produtos_no_pote_label = tk.Label(produtos_no_pote_frame, text="Açaí No pote", font=("Arial", 14))
    produtos_no_pote_label.pack()

    for produto in produtos_no_pote:
        tipo_produto = produto['tipo']
        valor_produto = produto['valor']

        botão_produto = tk.Button(produtos_no_pote_frame, text=f"{tipo_produto} - R$ {valor_produto:.2f}",
                                  command=lambda tipo_produto=tipo_produto, valor_produto=valor_produto:
                                  adicionar_pedido2(tipo_produto, valor_produto))
        botão_produto.pack()

    # COMPLEMENTOS
    complementos_frame = tk.Frame(janela_mesa)
    complementos_frame.pack(pady=20, side='left', anchor="n")

    complementos_label = tk.Label(complementos_frame, text="Complementos", font=("Arial", 14))
    complementos_label.pack()

    for produto in complementos:
        tipo_produto = produto['tipo']
        valor_produto = produto['valor']

        botão_produto = tk.Button(complementos_frame, text=f"{tipo_produto} - R$ {valor_produto:.2f}",
                                  command=lambda tipo_produto=tipo_produto, valor_produto=valor_produto:
                                  adicionar_pedido2(tipo_produto, valor_produto))
        botão_produto.pack()
    # ADICIONAIS

    adicionais_frame = tk.Frame(janela_mesa)
    adicionais_frame.pack(pady=20, side='left', anchor="n")

    adicionais_label = tk.Label(adicionais_frame, text="Adicionais", font=("Arial", 14))
    adicionais_label.pack()

    for produto in adicionais:
        tipo_produto = produto['tipo']
        valor_produto = produto['valor']

        botão_produto = tk.Button(adicionais_frame, text=f"{tipo_produto} - R$ {valor_produto:.2f}",
                                  command=lambda tipo_produto=tipo_produto, valor_produto=valor_produto:
                                  adicionar_pedido2(tipo_produto, valor_produto))
        botão_produto.pack()


    # BEBIDAS
    produtos_bebidas_frame = tk.Frame(janela_mesa)
    produtos_bebidas_frame.pack(pady=20, side="left", anchor='n')

    produtos_bebidas_label = tk.Label(produtos_bebidas_frame, text="Bebidas", font=("Arial", 14))
    produtos_bebidas_label.pack()

    for produto in produtos_bebidas:
        tipo_produto = produto['tipo']
        valor_produto = produto['valor']

        botão_produto = tk.Button(produtos_bebidas_frame, text=f"{tipo_produto} - R$ {valor_produto:.2f}",
                                  command=lambda tipo_produto=tipo_produto, valor_produto=valor_produto:
                                  adicionar_pedido2(tipo_produto, valor_produto))
        botão_produto.pack()

    principal_frame.pack(fill="both", expand=True)

    # Configure a barra de rolagem para controlar a visualização do frame principal


# JANELA RELATORIO
def abrir_janela_relatorio():
    janela_relatorio = tk.Toplevel(janela)
    janela_relatorio.title("Relatório de Vendas")

    tabela_relatorio_frame = tk.Frame(janela_relatorio)
    tabela_relatorio_frame.pack(pady=20)

    tabela_relatorio = ttk.Treeview(tabela_relatorio_frame, columns=(1, 2, 3, 4), show="headings", height=15)
    tabela_relatorio.pack(side="left")
    tabela_relatorio.heading(1, text="Mesa")
    tabela_relatorio.heading(2, text="Pedido")
    tabela_relatorio.heading(3, text="Valor")
    tabela_relatorio.heading(4, text="Forma de Pagamento")

    scrollbar = ttk.Scrollbar(tabela_relatorio_frame, orient="vertical", command=tabela_relatorio.yview)
    scrollbar.pack(side="right", fill="y")

    tabela_relatorio.configure(yscrollcommand=scrollbar.set)
    scrollbar.configure(command=tabela_relatorio.yview)

    for venda in vendas:
        mesa_numero = vendas.index(venda) + 1
        for pedido in venda['pedidos']:
            tipo_pedido = pedido['tipo']
            valor_pedido = pedido['valor']
            forma_pagamento = venda['forma_pagamento']
            tabela_relatorio.insert("", "end",
                                    values=(mesa_numero, tipo_pedido, f"R$ {valor_pedido:.2f}", forma_pagamento))

    def atualizar_total_caixam():
        total_caixam = caixa_inicial
        for mesa in vendas:
            forma_pagamento = mesa['forma_pagamento']
            if forma_pagamento == 'Dinheiro':
                valor_venda = mesa['valor_total']
                total_caixam += valor_venda
                valor_dindin.append(total_caixam)
        total_caixam_label.config(text=f"Total caixa físico: R$ {total_caixam / 2:.2f}")

    total_caixa_label = tk.Label(janela_relatorio,
                                 text=f"Total: R$ {sum(mesa['valor_total'] / 2 for mesa in vendas):.2f}",
                                 font=("Arial", 12))
    total_caixa_label.pack(pady=20)

    total_caixam_label = tk.Label(janela_relatorio, text="", font=("Arial", 12))
    total_caixam_label.pack(pady=20)

    valor_credito.clear()
    valor_debito.clear()
    valor_pix.clear()
    valor_dindin.clear()

    atualizar_total_caixam()  # Chamar a função para atualizar o total_caixam_label
    credito()
    pix()
    debito()




def carregar_valores_existentes():
    valores_existentes = {}
    data_resetada2 = date.today().strftime("%m-%Y")
    nome_arquivo5 = f"relatorio_formas_pagamento{data_resetada2}.xlsx"
    nome_pag = os.path.join(pasta_mes_ano, nome_arquivo5)
    try:
        planilha = openpyxl.load_workbook(nome_pag)
        planilha_ativa = planilha.active

        for row in planilha_ativa.iter_rows(min_row=2, values_only=True):
            forma_pagamento = row[0]  # Coluna A (índice 0)
            valor_venda = row[1]  # Coluna B (índice 1)
            valores_existentes[forma_pagamento] = valor_venda

        planilha.close()

    except FileNotFoundError:
        pass

    return valores_existentes


def salvar_valores_na_planilha(valores):
    planilha_formas_pagamento = openpyxl.Workbook()
    planilha_ativa_formas_pagamento = planilha_formas_pagamento.active

    planilha_ativa.append(["Forma de Pagamento", "Valor da Venda"])
    planilha_ativa_formas_pagamento.append(["Forma de Pagamento", "Valor da Venda"])

    for forma_pagamento, valor_venda in valores.items():
        planilha_ativa.append([forma_pagamento, valor_venda])
        planilha_ativa_formas_pagamento.append([forma_pagamento, valor_venda])

    data_resetada2 = date.today().strftime("%m-%Y")
    nome_arquivo5 = f"relatorio_formas_pagamento{data_resetada2}.xlsx"
    nome_pag = os.path.join(pasta_mes_ano, nome_arquivo5)
    planilha_formas_pagamento.save(nome_pag)
    planilha_formas_pagamento.close()


def criar_planilha_separada(vendas):
    valores_existentes = carregar_valores_existentes()

    totais_vendas = {}

    for venda in vendas:
        for pedido in venda['pedidos']:
            forma_pagamento = venda['forma_pagamento']
            valor_pedido = pedido['valor']

            if forma_pagamento not in totais_vendas:
                totais_vendas[forma_pagamento] = valor_pedido
            else:
                totais_vendas[forma_pagamento] += valor_pedido

    for forma_pagamento, valor_venda in totais_vendas.items():
        if forma_pagamento in valores_existentes:
            valores_existentes[forma_pagamento] += valor_venda
        else:
            valores_existentes[forma_pagamento] = valor_venda

    salvar_valores_na_planilha(valores_existentes)

    # Atualizar o valor total na planilha "relatorio_formas_pagamento.xlsx"
    data_resetada2 = date.today().strftime("%m-%Y")
    nome_arquivo5 = f"relatorio_formas_pagamento{data_resetada2}.xlsx"
    nome_pag = os.path.join(pasta_mes_ano, nome_arquivo5)
    valor_total = sum(valores_existentes.values())
    planilha_total = openpyxl.load_workbook(nome_pag)
    planilha_total.active['A1'] = "Valor Total"
    planilha_total.active['B1'] = valor_total
    nome_pag = os.path.join(pasta_mes_ano, nome_arquivo5)
    planilha_total.save(nome_pag)

def carregar_valores_existentes2():
    valores_existentes = {}
    caminho_arquivo = os.path.join(pasta_documentos, pasta_mes_atual, nome_arquivo)
    nome_pag = os.path.join(caminho_arquivo)
    try:
        planilha = openpyxl.load_workbook(nome_pag)
        planilha_ativa = planilha.active

        for forma_pagamento, valor_venda in planilha_ativa.iter_rows(min_row=2, values_only=True, max_col=2):
            valores_existentes[forma_pagamento] = valor_venda

        planilha.close()

    except FileNotFoundError:
        pass

    return valores_existentes


def salvar_valores_na_planilha2(valores):
    nome_pag = os.path.join(caminho_arquivo)
    planilha_formas_pagamento = openpyxl.load_workbook(nome_pag)
    planilha_ativa_formas_pagamento = planilha_formas_pagamento.active


    for forma_pagamento, valor_venda in valores.items():
        if forma_pagamento == "Dinheiro":
            planilha_ativa_formas_pagamento.append([forma_pagamento, valor_venda])
        if forma_pagamento == "Débito":
            planilha_ativa_formas_pagamento.append([forma_pagamento, valor_venda])
        if forma_pagamento == "Crédito":
            planilha_ativa_formas_pagamento.append([forma_pagamento, valor_venda])
        if forma_pagamento == "Pix":
            planilha_ativa_formas_pagamento.append([forma_pagamento, valor_venda])

    planilha_formas_pagamento.save(nome_pag)
    planilha_formas_pagamento.close()


def criar_planilha_separada2(vendas):
    valores_existentes = carregar_valores_existentes2()

    totais_vendas = {}

    for venda in vendas:
        for pedido in venda['pedidos']:
            forma_pagamento = venda['forma_pagamento']
            valor_pedido = pedido['valor']

            if forma_pagamento not in totais_vendas:
                totais_vendas[forma_pagamento] = valor_pedido
            else:
                totais_vendas[forma_pagamento] += valor_pedido

    for forma_pagamento, valor_venda in totais_vendas.items():
        if forma_pagamento in valores_existentes:
            valores_existentes[forma_pagamento] += valor_venda
        else:
            valores_existentes[forma_pagamento] = valor_venda

    salvar_valores_na_planilha2(valores_existentes)

    # Atualizar o valor total na planilha "relatorio_formas_pagamento.xlsx"


    nome_pag = os.path.join(caminho_arquivo)
    planilha_total = openpyxl.load_workbook(nome_pag)
    nome_pag = os.path.join(caminho_arquivo)
    planilha_total.save(nome_pag)

def enviar_arquivos_por_email2(diretorio, destinatario):
    data_resetada2 = date.today().strftime("%m-%Y")
    nome_arquivo5 = f"Relatório Mensal{data_resetada2}.xlsx"
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = destinatario
    mail.Subject = f'Total Mensal (Qe 30) {data_resetada2}'
    mail.HTMLBody = f'''
    <p>Prezado, Rafael</p>

    <p>Segue o valor mensal total do mês {data_resetada2}.</p>

    <p>Att.,</p>
    <p>Extremo sabor.</p>
    '''

    for nome_arquivo5 in os.listdir(pasta_mes_ano):
        caminho_arquivo = os.path.join(pasta_mes_ano, nome_arquivo5)
        if os.path.isfile(caminho_arquivo):  # Verifica se é um arquivo
            attachment = os.path.abspath(caminho_arquivo)
            mail.Attachments.Add(attachment)

    mail.Send()
    print('Email enviado com sucesso.')


import calendar
import datetime


def email():
    global fechando
    diretorio_base = os.path.join(os.path.expanduser("~"), "Documents", "ExtremoSabor", "Nota Fiscal")

    def atualizar_planilha():
        criar_planilha_separada(vendas)
        planilha_diaria_atualizar_valores()
        criar_planilha_separada2(vendas)
    atualizar_planilha()
    # Obter a data atual
    data_atual = datetime.date.today()
    ultimo_dia_mes = calendar.monthrange(data_atual.year, data_atual.month)[1]
    nome_pasta = 'Nota Fiscal-' + data_atual.strftime("%Y-%m")

    # Diretório completo para salvar o relatório
    diretorio = os.path.join(diretorio_base, nome_pasta)
    nome_pasta2 = data_atual.strftime("%d-%m-%Y")
    diretorio2 = os.path.join(diretorio_base, nome_pasta, nome_pasta2)
    resposta = messagebox.askquestion("Fechar Programa",
                                      "Deseja enviar os arquivos por email antes de fechar o programa?")
    if resposta == 'yes':
        enviar_arquivos_por_email(diretorio2, 'rafaelhenrycassiano23@gmail.com')
        pasta_documentos = os.path.join(os.path.expanduser("~"), "Documents", "ExtremoSabor", "Relatório")
        nome_pasta = 'Relatório-' + date.today().strftime("%Y-%h")
        pasta_mes_atual = os.path.join(pasta_documentos, nome_pasta)
        data_resetada = date.today().strftime("%d-%m-%Y")
        nome_arquivo = f"Relatorio_{data_resetada}.xlsx"
        caminho_arquivo = os.path.join(pasta_documentos, pasta_mes_atual, nome_arquivo)
        # importar a base de dados
        tabela_vendas = pd.read_excel(caminho_arquivo)

        # visualizar a base de dados
        pd.set_option('display.max_columns', None)
        print(tabela_vendas)

        # enviar um email com o relatório
        outlook = win32.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = 'rafaelhenrycassiano23@gmail.com'
        mail.Subject = f'Relatório de Vendas (Qe 30) do dia {data_resetada}'
        mail.HTMLBody = f'''
                <p>Prezado, Rafael</p>

                <p>Segue o Relatório de Vendas do dia {data_resetada}, onde NaN representam os espaços vazios da planilha.</p>



                <p>Quantidade Vendida:</p>
                {tabela_vendas.to_html()}


                <p>Att.,</p>
                <p>Extremo sabor.</p>
                '''

        mail.Send()
        if data_atual == ultimo_dia_mes:
            resposta2 = messagebox.askquestion("Fechar Programa",
                                               "Hoje é o último dia do mês, deseja enviar os arquivos mensais por email antes de fechar o programa?")
            if resposta2 == 'yes':
                data_resetada8 = date.today().strftime("%m-%Y")
                nome_arquivo2 = f"relatorio_formas_pagamento{data_resetada8}.xlsx"
                caminho_arquivo2 = os.path.join(pasta_mes_ano, nome_arquivo2)
                # importar a base de dados
                data_resetada3 = date.today().strftime("%m-%Y")
                mensal = pd.read_excel(caminho_arquivo2)

                # visualizar a base de dados
                pd.set_option('display.max_columns', None)
                print(mensal)

                # enviar um email com o relatório
                outlook = win32.Dispatch('outlook.application')
                mail = outlook.CreateItem(0)
                mail.To = 'rafaelhenrycassiano23@gmail.com'
                mail.Subject = f'Seguem os valores mensais/pagamentos (Qe 30). {data_resetada3}'
                mail.HTMLBody = f'''
                                <p>Prezado, Rafael</p>

                                <p>Seguem os valores mensais.</p>



                                <p>Quantidade Vendida:</p>
                                {mensal.to_html()}
                                <p>Att.,</p>
                        <p>Extremo sabor.</p>
                        '''

                mail.Send()
                ####################### EMAIL MENSAL
                data_resetada5 = date.today().strftime("%m-%Y")
                nome_arquivo4 = f"Relatório Mensal{data_resetada5}.xlsx"
                caminho_arquivo5 = os.path.join(pasta_relatorio, pasta_mes_ano, nome_arquivo4)
                # importar a base de dados
                tabela_mensal = pd.read_excel(caminho_arquivo5)

                # visualizar a base de dados
                pd.set_option('display.max_columns', None)
                print(tabela_mensal)

                # enviar um email com o relatório
                outlook = win32.Dispatch('outlook.application')
                mail = outlook.CreateItem(0)
                mail.To = 'rafaelhenrycassiano23@gmail.com'
                mail.Subject = f'Relatório Mensal (Qe 30) {data_resetada5}.'
                mail.HTMLBody = f'''
                                <p>Prezado, Rafael</p>

                                <p>Segue o Relatório Mensal {data_resetada5}.</p>



                                <p>Quantidade Vendida:</p>
                                {tabela_mensal.to_html()}


                                <p>Att.,</p>
                                <p>Extremo sabor.</p>
                                '''

                mail.Send()
            else:
                janela.destroy()

            print('Email Enviado')

            janela.destroy()

        print('Email Enviado')




        janela.destroy()
    else:
        janela.destroy()


abrir_mesa_button = tk.Button(janela, text="Abrir Mesa", command=abrir_janela_mesa)
abrir_mesa_button.pack(pady=20)

abrir_relatorio_button = tk.Button(janela, text="Abrir Relatório", command=abrir_janela_relatorio)
abrir_relatorio_button.pack()

fechando = False
janela.protocol("WM_DELETE_WINDOW", email)
janela.mainloop()
